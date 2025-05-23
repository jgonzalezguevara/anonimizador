#!/usr/bin/env python3
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import re
import fitz  # PyMuPDF
from docx import Document
from openpyxl import load_workbook

# Diccionario de patrones genéricos activables desde anonimizables.txt
PATRONES_GENERICOS = {
    "EMAIL": r"\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-z]{2,}\b",
    "URL": r"https?://\S+",
    "WEB": r"www\.[a-zA-Z0-9.-]+\.[a-z]{2,}",
    "DNI": r"\b\d{8}[A-Za-z]\b",
    "NIF": r"\b[A-Z]\d{8}\b",
    "CIF": r"\b[A-Z]\d{8}\b",
    "VALOR_ECONÓMICO": r"\d{1,3}(\.\d{3})*(,\d{2})?\s?(euros|EUR|€)?",
    "CÓDIGO_NUTS": r"\bES\d{3}\b",
    "CÓDIGO_CPV": r"\b\d{8}\b",
    "NOMBRE_PERSONA": r"[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+\s+[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+",
    "DIRECCIÓN": r"(?i)(Calle|C/|Avenida|Av\.|Plaza|Pza\.)\s+[\w\s]+,?\s*\d+",
    "DOMINIO_INSTITUCIONAL": r"(?i)\b(adif|boe|contrataciondelestado)\.es\b",
    "NUMERO_IDENTIFICACION": r"\bQ\d{7}[A-Z]\b",
    "CORREO_ELECTRÓNICO": r"(?i)(Correo electrónico|Email):?\s*[^\s]+@[^\s]+",
    "VALOR_ESTIMADO": r"(?i)Valor estimado:?\s*\d{1,3}(\.\d{3})*(,\d{2})?\s?(euros|EUR|€)?",
    "PERFIL_COMPRADOR": r"(?i)perfil de comprador[:\s]+.*",
    "DIRECCIÓN_PRINCIPAL": r"(?i)dirección principal[:\s]+.*",
    "PORTAL_CONTRATACIÓN": r"(?i)contrataciondelestado\.es"
}


def anonimizar_texto(texto, claves_activas):
    for clave in claves_activas:
        patron = PATRONES_GENERICOS.get(clave)
        if patron:
            texto = re.sub(patron, '***', texto, flags=re.IGNORECASE)
    return texto


def cargar_claves_anonimizables():
    try:
        with open("anonimizables.txt", "r", encoding="utf-8") as f:
            claves = [line.strip().upper() for line in f if line.strip() and not line.startswith('#')]
            return claves
    except FileNotFoundError:
        messagebox.showwarning("Aviso", "No se encontró 'anonimizables.txt'. Se anonimizarán solo patrones genéricos.")
        return []


def obtener_ruta_salida(filepath, output_dir):
    nombre_archivo = os.path.basename(filepath)
    nombre, extension = os.path.splitext(nombre_archivo)
    return os.path.join(output_dir, f"{nombre}_anonimizado{extension}")


def anonimizar_docx(filepath, claves, output_dir):
    doc = Document(filepath)
    for para in doc.paragraphs:
        texto_completo = "".join(run.text for run in para.runs)
        texto_anon = anonimizar_texto(texto_completo, claves)
        if texto_completo != texto_anon:
            for run in para.runs:
                run.text = ""
            if para.runs:
                para.runs[0].text = texto_anon

    if "ELIMINAR_IMAGENES" in claves:
        for shape in doc.inline_shapes:
            shape._inline.graphic.graphicData._element.clear()

    doc.save(obtener_ruta_salida(filepath, output_dir))
    print(f"[DOCX] Anonimizado: {os.path.basename(filepath)}")


def anonimizar_xlsx(filepath, claves, output_dir):
    wb = load_workbook(filepath)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    texto_anon = anonimizar_texto(cell.value, claves)
                    if cell.value != texto_anon:
                        cell.value = texto_anon
    wb.save(obtener_ruta_salida(filepath, output_dir))
    print(f"[XLSX] Anonimizado: {os.path.basename(filepath)}")


def anonimizar_pdf(filepath, claves, output_dir):
    doc = fitz.open(filepath)
    patrones = [PATRONES_GENERICOS[c] for c in claves if c in PATRONES_GENERICOS]

    for page in doc:
        texto_original = page.get_text("text")

        for patron in patrones:
            matches = re.finditer(patron, texto_original, flags=re.IGNORECASE)
            for match in matches:
                fragmento = match.group()
                areas = page.search_for(fragmento)
                for rect in areas:
                    page.add_redact_annot(rect, fill=(1, 1, 1))

        if "ELIMINAR_IMAGENES" in claves:
            for img in page.get_images(full=True):
                xref = img[0]
                rects = page.get_image_rects(xref)
                for rect in rects:
                    page.add_redact_annot(rect, fill=(1, 1, 1))

        page.apply_redactions()

    doc.save(obtener_ruta_salida(filepath, output_dir))
    print(f"[PDF] Anonimizado: {os.path.basename(filepath)}")


def seleccionar_archivo():
    filepath = filedialog.askopenfilename(filetypes=[
        ("Documentos soportados", "*.docx *.xlsx *.pdf")
    ])
    if not filepath:
        return

    output_dir = filedialog.askdirectory(title="Selecciona carpeta de destino")
    if not output_dir:
        return

    try:
        claves = cargar_claves_anonimizables()
        ext = filepath.lower()

        if ext.endswith(".docx"):
            anonimizar_docx(filepath, claves, output_dir)
        elif ext.endswith(".xlsx"):
            anonimizar_xlsx(filepath, claves, output_dir)
        elif ext.endswith(".pdf"):
            anonimizar_pdf(filepath, claves, output_dir)
        else:
            messagebox.showerror("Error", "Tipo de archivo no soportado.")
            return

        messagebox.showinfo("Éxito", "Archivo anonimizado correctamente.")

    except Exception as e:
        messagebox.showerror("Error", f"Se produjo un error: {str(e)}")


def main():
    root = tk.Tk()
    root.withdraw()
    seleccionar_archivo()


if __name__ == "__main__":
    main()

